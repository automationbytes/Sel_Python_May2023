'''

Openpyxl - external lib
light weight
closest match to apache poi

Pandas - dataframes

'''
import openpyxl
def readExcel(sheetname,label,header):
    workbook = openpyxl.load_workbook("Data/Datasheet1.xlsx")
    sheet = workbook[sheetname]

    rows = sheet.max_row
    col = sheet.max_column

    Label = []
    Header = []
    #
    # for i in range(1,rows+1):
    #     Label.append(sheet.cell(i,1).value)
    # print(Label)
    #
    #
    # for j in range(1,col+1):
    #     Header.append(sheet.cell(1,j).value)
    # print(Header)
    for i in range(1, rows + 1):
        Label.append(sheet.cell(i, 1).value)
        if (Label[i - 1] == label):
            for j in range(1, col + 1):
                Header.append(sheet.cell(1, j).value)
                if (Header[j - 1] == header):
                    val = sheet.cell(i, j).value
                    break
            break
    return (val)

print(readExcel("Priya","Amazon","URL"))